import requests
from bs4 import BeautifulSoup
import json
import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import random

def spider(ip_param):
    url = "http://ip.taobao.com/service/getIpInfo.php?ip={}".format(ip_param)
    try:
        wb_data = requests.get(url)
    except Exception:
        time.sleep(3)
        wb_data = requests.get(url)
    soup = BeautifulSoup(wb_data.text, "lxml")
    text_json = json.loads(soup.text)
    if text_json['code'] == 0:
        print(text_json)
        country = text_json['data']['country']
        region = text_json['data']['region']
        city = text_json['data']['city']
        isp = text_json['data']['isp']

        final_dict = {
            "country": country,
            'region': region,
            'city': city,
            'isp': isp
        }

        final_list = [country, region, city, isp]
        return final_list
    else:
        return ['', '', '', '']


def read_excel(file_name):
    work_book = xlrd.open_workbook(r"{}".format(file_name))
    first_sheet = work_book.sheet_by_index(0)
    ret = []
    for i in range(1, first_sheet.nrows):
        row = first_sheet.row_values(i)
        ret.append(row)
    return ret

def find_line_numbers(file_name):
    work_book = xlrd.open_workbook(r"{}".format(file_name))
    row_numbers = work_book.sheet_by_index(0).nrows
    return row_numbers


def openpyxl_write():
    wb = Workbook()
    ws = wb.active
    ws.title = "New Title"
    d = ws.cell(row=4, column=2)
    d.value = "asd"
    wb.save("test.xlsx")

if __name__ == '__main__':
    line_numbers = find_line_numbers("test.xlsx")
    excel_list = read_excel("c_ip.xlsx")
    huge_list = []
    wb = load_workbook('test.xlsx')
    for w_index, row in enumerate(excel_list[line_numbers:]):
        ts = random.random()*5
        time.sleep(1+ts)
        ip = row[1]
        count = row[2]
        temp_list = spider(ip)
        temp_list.insert(0, count)
        temp_list.insert(0, ip)
        print(temp_list)

        ws= wb.get_sheet_by_name("New Title")
        for index, item in enumerate(temp_list):
            d = ws.cell(row=w_index+line_numbers+1, column=index + 1)
            d.value = item
        wb.save('test.xlsx')
    wb.close()
